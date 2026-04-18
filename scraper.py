"""
E-Commerce Price Monitor
- Scrapes product prices from books.toscrape.com (public demo site)
- Stores history in SQLite
- Detects price drops and exports to CSV/Excel
"""
import sqlite3, csv, time, re, os
from datetime import datetime
from urllib.request import urlopen, Request
from urllib.error import URLError
from html.parser import HTMLParser

DB_PATH  = "prices.db"
BASE_URL = "http://books.toscrape.com"
MAX_PAGES = 5          # scrape first 5 pages (~100 books)
DELAY     = 1.0        # seconds between requests

RATING_MAP = {"One":1,"Two":2,"Three":3,"Four":4,"Five":5}


# ── DB setup ───────────────────────────────────────────────────────────────
def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS prices (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            title     TEXT,
            price     REAL,
            rating    INTEGER,
            in_stock  INTEGER,
            url       TEXT,
            scraped   TEXT
        )
    """)
    con.commit()
    return con


# ── Minimal HTML parser ────────────────────────────────────────────────────
class BookParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.books  = []
        self._cur   = {}
        self._in_title = self._in_price = self._in_stock = False

    def handle_starttag(self, tag, attrs):
        attrs = dict(attrs)
        if tag == "article" and "product_pod" in attrs.get("class",""):
            self._cur = {}
        if tag == "h3":
            pass
        if tag == "a" and self._cur is not None and "title" in attrs:
            self._cur["title"] = attrs["title"]
            self._cur["url"]   = attrs.get("href","")
        if tag == "p" and "price_color" in attrs.get("class",""):
            self._in_price = True
        if tag == "p" and "star-rating" in attrs.get("class",""):
            word = attrs.get("class","").replace("star-rating","").strip()
            self._cur["rating"] = RATING_MAP.get(word, 0)
        if tag == "p" and "instock" in attrs.get("class",""):
            self._in_stock = True

    def handle_data(self, data):
        if self._in_price:
            price_str = re.sub(r"[^\d.]", "", data)
            if price_str:
                self._cur["price"] = float(price_str)
            self._in_price = False
        if self._in_stock:
            self._cur["in_stock"] = 1 if "In stock" in data else 0
            self._in_stock = False

    def handle_endtag(self, tag):
        if tag == "article" and self._cur.get("title"):
            self.books.append(self._cur)
            self._cur = {}


def fetch(url: str) -> str:
    req = Request(url, headers={"User-Agent": "Mozilla/5.0 (price-monitor-demo/1.0)"})
    with urlopen(req, timeout=10) as r:
        return r.read().decode("utf-8", errors="replace")


# ── Scraper ────────────────────────────────────────────────────────────────
def scrape_all() -> list[dict]:
    all_books = []
    for page in range(1, MAX_PAGES + 1):
        url = f"{BASE_URL}/catalogue/page-{page}.html"
        print(f"  scraping page {page}: {url}")
        try:
            html = fetch(url)
        except URLError as e:
            print(f"  [err] {e}")
            break
        parser = BookParser()
        parser.feed(html)
        print(f"    found {len(parser.books)} books")
        all_books.extend(parser.books)
        time.sleep(DELAY)
    return all_books


def save_to_db(con: sqlite3.Connection, books: list[dict]):
    ts = datetime.now().isoformat(timespec="seconds")
    con.executemany(
        "INSERT INTO prices (title,price,rating,in_stock,url,scraped) VALUES (?,?,?,?,?,?)",
        [(b.get("title",""), b.get("price",0), b.get("rating",0),
          b.get("in_stock",1), b.get("url",""), ts) for b in books]
    )
    con.commit()
    print(f"  saved {len(books)} records to DB")


# ── Price-drop detection ───────────────────────────────────────────────────
def find_price_drops(con: sqlite3.Connection) -> list[dict]:
    """Compare latest price vs previous price for each title."""
    rows = con.execute("""
        SELECT title,
               MAX(CASE WHEN rn=1 THEN price END) AS latest,
               MAX(CASE WHEN rn=2 THEN price END) AS previous
        FROM (
            SELECT title, price,
                   ROW_NUMBER() OVER (PARTITION BY title ORDER BY scraped DESC) AS rn
            FROM prices
        )
        WHERE rn <= 2
        GROUP BY title
        HAVING previous IS NOT NULL AND latest < previous
    """).fetchall()
    return [{"title":r[0],"latest":r[1],"previous":r[2],
             "drop": round(r[2]-r[1],2),
             "pct":  round((r[2]-r[1])/r[2]*100,1)} for r in rows]


# ── Export ─────────────────────────────────────────────────────────────────
def export_csv(con: sqlite3.Connection, path: str = "latest_prices.csv"):
    rows = con.execute(
        "SELECT title,price,rating,in_stock,url,scraped FROM prices "
        "WHERE scraped = (SELECT MAX(scraped) FROM prices)"
    ).fetchall()
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Title","Price(£)","Rating","In Stock","URL","Scraped"])
        w.writerows(rows)
    print(f"  exported {len(rows)} rows → {path}")


def export_excel(con: sqlite3.Connection, path: str = "price_report.xlsx"):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("  openpyxl not installed, skipping Excel export")
        return

    rows = con.execute(
        "SELECT title,price,rating,in_stock,url,scraped FROM prices "
        "WHERE scraped = (SELECT MAX(scraped) FROM prices) ORDER BY price ASC"
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price Report"

    headers = ["Title","Price (£)","Rating","In Stock","URL","Scraped At"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            if ci == 4:   # In Stock
                c.value = "Yes" if val else "No"
            if ci == 2:   # Price — color cheap books green
                if isinstance(val, (int,float)) and val < 20:
                    c.fill = PatternFill("solid", fgColor="C6EFCE")

    ws.auto_filter.ref = ws.dimensions
    ws.column_dimensions["A"].width = 50
    wb.save(path)
    print(f"  Excel report → {path}")


# ── Main ───────────────────────────────────────────────────────────────────
def main():
    print("=== E-Commerce Price Monitor ===")
    print(f"Target: {BASE_URL}  |  Pages: {MAX_PAGES}")

    con = init_db()

    print("\n[1] Scraping prices...")
    books = scrape_all()
    print(f"    Total scraped: {len(books)} books")

    print("\n[2] Saving to database...")
    save_to_db(con, books)

    print("\n[3] Checking price drops...")
    drops = find_price_drops(con)
    if drops:
        print(f"    {len(drops)} price drop(s) detected:")
        for d in drops[:5]:
            print(f"    ↓ {d['title'][:50]:50s}  £{d['previous']} → £{d['latest']}  (-{d['pct']}%)")
    else:
        print("    No drops yet (need 2+ scrape runs to compare)")

    print("\n[4] Exporting results...")
    export_csv(con)
    export_excel(con)

    print("\n✅ Done!  Run again later to detect price changes.")
    con.close()


if __name__ == "__main__":
    main()
